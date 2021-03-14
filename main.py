from __future__ import annotations

import os

from calendar import monthrange

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from workalendar.europe import Poland

from employee import Employee
from functions import find_first
from gender import Gender
from summary import MonthlySummary
from writer import Writer


def get_employees(workbook: Workbook) -> list[Employee]:
    return [Employee(row[1], row[0], Gender.from_string(row[2]))
            for row
            in workbook.active.iter_rows(min_row=2, max_col=3, values_only=True)]


WORKING_DIRECTORY: str = "C:\\Users\\User\\Desktop\\ROK 2021"
EMPLOYEES_FILE: str = "Pracownicy.xlsx"

months: list[str] = ["Styczeń", "Luty", "Marzec",
                     "Kwiecień", "Maj", "Czerwiec",
                     "Lipiec", "Sierpień", "Wrzesień",
                     "Październik", "Listopad", "Grudzień"]
year: int = 2021

employees: list[Employee] = get_employees(load_workbook(os.path.join(WORKING_DIRECTORY, EMPLOYEES_FILE)))
dirs = [file for file in os.listdir(WORKING_DIRECTORY) if os.path.isdir(os.path.join(WORKING_DIRECTORY, file))]

for month_no, month in enumerate(months):
    # print(month)
    dir_index = find_first(dirs, lambda x: month.lower() in x.lower())
    directory = dirs[dir_index] if dir_index != -1 else None
    for i, employee in enumerate(employees):
        # print(f"  {i}. {employee.name} {employee.last_name}")
        if directory:
            summary_file = employee.find_summary(os.path.join(WORKING_DIRECTORY, directory))
            if summary_file:
                try:
                    employee.monthly_summaries[month] = MonthlySummary.get_from_worksheet(
                        load_workbook(summary_file)
                        .active,
                        width=monthrange(2021, month_no + 1)[1]
                    )
                    continue
                except Exception as e:
                    employee.monthly_summaries[month] = MonthlySummary.empty()
        employee.monthly_summaries[month] = MonthlySummary.empty()

extra_holidays: list = []

writer = Writer(f"Ewidencja godzin {year}.xlsx", year, employees[:10], Poland(), extra_holidays, months)
writer.create()
